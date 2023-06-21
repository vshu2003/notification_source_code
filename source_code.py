import time
import smtplib
import openpyxl
from tkinter import Tk, Label, Button, messagebox
from tkinter import ttk
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from cryptography.fernet import Fernet
import schedule

# Function to encrypt the LinkedIn credentials
def encrypt_credentials(key, username, password):
    cipher_suite = Fernet(key)
    encrypted_username = cipher_suite.encrypt(username.encode())
    encrypted_password = cipher_suite.encrypt(password.encode())
    return encrypted_username, encrypted_password

# Function to decrypt the LinkedIn credentials
def decrypt_credentials(key, encrypted_username, encrypted_password):
    cipher_suite = Fernet(key)
    decrypted_username = cipher_suite.decrypt(encrypted_username).decode()
    decrypted_password = cipher_suite.decrypt(encrypted_password).decode()
    return decrypted_username, decrypted_password

# Function to log in to LinkedIn and retrieve unread notification count
def get_unread_notification_count(username, password):
    driver = webdriver.Chrome()
    driver.get('https://www.linkedin.com/login')
    driver.find_element_by_id('username').send_keys(username)
    driver.find_element_by_id('password').send_keys(password)
    driver.find_element_by_id('password').send_keys(Keys.ENTER)
    time.sleep(5)
    notification_count = driver.find_element_by_css_selector('span.nav-item__count').text
    driver.quit()
    return notification_count

# Function to log in to LinkedIn and retrieve unread message count
def get_unread_message_count(username, password):
    driver = webdriver.Chrome()
    driver.get('https://www.linkedin.com/login')
    driver.find_element_by_id('username').send_keys(username)
    driver.find_element_by_id('password').send_keys(password)
    driver.find_element_by_id('password').send_keys(Keys.ENTER)
    time.sleep(5)
    message_count = driver.find_element_by_css_selector('span.msg-overlay-bubble-header__count').text
    driver.quit()
    return message_count

# Function to send email notification
def send_email_notification(sender_email, sender_password, recipient_email, message):
    with smtplib.SMTP('smtp.gmail.com', 587) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, recipient_email, message)

# Function to handle the button click event
def on_notify_click():
    # Set up the encryption key
    key = b'your_secret_key'

    # Encrypt the LinkedIn credentials
    encrypted_linkedin_username, encrypted_linkedin_password = encrypt_credentials(key, 'your_linkedin_username', 'your_linkedin_password')

    # Decrypt the LinkedIn credentials
    linkedin_username, linkedin_password = decrypt_credentials(key, encrypted_linkedin_username, encrypted_linkedin_password)

    # Set up the email credentials
    sender_email = 'your_sender_email@gmail.com'
    sender_password = 'your_sender_email_password'
    recipient_email = 'your_recipient_email@gmail.com'

    # Get the current unread counts from LinkedIn
    notification_count = get_unread_notification_count(linkedin_username, linkedin_password)
    message_count = get_unread_message_count(linkedin_username, linkedin_password)

    # Load the previous counts from Excel
    excel_file = 'linkedin_counts.xlsx'
    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.active
    previous_notification_count = worksheet.cell(row=2, column=1).value
    previous_message_count = worksheet.cell(row=2, column=2).value

    # Save the current counts to Excel for future usage
    worksheet.cell(row=2, column=1).value = notification_count
    worksheet.cell(row=2, column=2).value = message_count
    workbook.save(excel_file)

    # Create the email body
    email_body = f'''
    <h2>LinkedIn Unread Counts</h2>
    <p>Number of Unread Messages: {message_count}</p>
    <p>Number of Unread Notifications: {notification_count}</p>
    <p>Comparison with Previous Data:</p>
    <p>Messages: {message_count} (Current) - {previous_message_count} (Previous)</p>
    <p>Notifications: {notification_count} (Current) - {previous_notification_count} (Previous)</p>
    '''

    # Create the email message
    message = f'''From: {sender_email}
    To: {recipient_email}
    Subject: LinkedIn Unread Counts

    {email_body}
    '''

    # Send email notification
    send_email_notification(sender_email, sender_password, recipient_email, message)
    messagebox.showinfo("Notification", "LinkedIn Unread Counts have been sent!")

# Function to schedule the notification every 3 hours
def schedule_notifications():
    schedule.every(3).hours.do(on_notify_click)

    while True:
        schedule.run_pending()
        time.sleep(1)

# Create the GUI window
def create_gui():
    window = Tk()
    window.title("LinkedIn Unread Counts Notification")
    window.geometry("400x200")
    window.configure(bg="white")

    # Create the title label
    title_label = Label(window, text="LinkedIn Unread Counts Notification", font=("Arial", 14), bg="white")
    title_label.pack(pady=20)

    # Create the notification button
    btn_notify = Button(window, text="Send Notification", command=on_notify_click, bg="lightblue", fg="black", font=("Arial", 12))
    btn_notify.pack(pady=10)

    # Run the GUI window
    window.mainloop()

# Run the GUI window
create_gui()

# Start scheduling the notifications
schedule_notifications()
